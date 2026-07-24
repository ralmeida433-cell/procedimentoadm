"""Leitura/escrita da planilha de controle (Excel) usada como fonte de dados.

Cada linha da planilha representa um PCD. As colunas de controle (Status,
Status Vista Inicial, Status Vista Final, e suas respectivas Mensagens)
são escritas de volta pelo sistema após cada etapa, para que a unidade
veja o resultado sem precisar abrir os logs.
"""
from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

# (chave_schema, cabeçalho, tipo) - tipo em {"texto", "data", "sim_nao"}
#
# Nota: cada estágio do processo (instauração, vista inicial, vista final)
# tem sua própria data de citação e sua própria contagem de folhas dos
# autos - a contagem cresce conforme documentos são juntados ao processo,
# por isso não são compartilhadas entre estágios.
DEFINICAO_COLUNAS = [
    ("reds", "REDS", "texto"),
    ("nome_sindicado", "Nome do Sindicado", "texto"),
    ("re_sindicado", "Número de Matrícula do Sindicado", "texto"),
    ("posto_graduacao_sindicado", "Posto/Graduação do Sindicado", "texto"),
    ("unidade_sindicado", "Unidade do Sindicado", "texto"),
    ("data_fato", "Data do Fato (dd/mm/aaaa)", "data"),
    ("cidade_fato", "Cidade do Fato", "texto"),
    ("hora_fato", "Hora do Fato", "texto"),
    ("resumo_fato", "Resumo do Fato", "texto"),
    ("nome_autoridade_processante", "Nome da Autoridade Processante", "texto"),
    ("posto_autoridade_processante", "Posto/Graduação da Autoridade Processante", "texto"),
    ("re_autoridade_processante", "Número de Matrícula da Autoridade Processante", "texto"),
    ("unidade_autoridade_processante", "Unidade da Autoridade Processante", "texto"),
    ("estavel_autoridade_processante", "Autoridade Processante é Estável? (S/N)", "sim_nao"),
    ("nome_autoridade_delegante", "Nome da Autoridade Delegante", "texto"),
    ("posto_autoridade_delegante", "Posto da Autoridade Delegante", "texto"),
    ("numero_processo", "Número do Processo", "texto"),
    ("numero_regiao_pm", "Nº Região PM", "texto"),
    ("numero_batalhao_pm", "Nº Batalhão PM", "texto"),
    ("numero_comunicacao_disciplinar", "Nº da Comunicação Disciplinar", "texto"),
    ("numero_folhas_comunicacao_disciplinar", "Nº de Folhas da Comunicação Disciplinar", "texto"),
    ("numero_folhas_escala_servico", "Nº de Folhas da Escala de Serviço Anexa", "texto"),
    ("local_fato", "Local do Fato", "texto"),
    ("tipificacao_cedm", "Tipificação CEDM (inciso/artigo)", "texto"),
    ("cidade_sede", "Cidade Sede da Unidade", "texto"),
    ("data_instauracao", "Data de Instauração (dd/mm/aaaa)", "data"),
    # Quem comunicou a transgressão (Comunicação Disciplinar)
    ("nome_comunicante", "Nome do Comunicante", "texto"),
    ("posto_comunicante", "Posto/Graduação do Comunicante", "texto"),
    ("re_comunicante", "Número de Matrícula do Comunicante", "texto"),
    ("unidade_comunicante", "Unidade do Comunicante", "texto"),
    ("data_comunicacao", "Data da Comunicação Disciplinar (dd/mm/aaaa)", "data"),
    # Testemunha da Comunicação Disciplinar (opcional)
    ("nome_testemunha", "Nome da Testemunha (Comunicação)", "texto"),
    ("posto_testemunha", "Posto/Graduação da Testemunha (Comunicação)", "texto"),
    ("re_testemunha", "Número de Matrícula da Testemunha (Comunicação)", "texto"),
    ("unidade_testemunha", "Unidade da Testemunha (Comunicação)", "texto"),
    ("bens_documentos_relacionados", "Bens/Documentos Relacionados", "texto"),
    ("parentesco_ou_inimizade", "Impedimento Declarado? (S/N)", "sim_nao"),
    ("observacoes_impedimento", "Observações sobre Impedimento", "texto"),
    ("prorrogado", "Prazo Prorrogado? (S/N)", "sim_nao"),
    ("data_conclusao_real", "Data de Conclusão Real (dd/mm/aaaa)", "data"),
    # Vista Inicial (defesa prévia)
    ("data_citacao", "Data da Citação Inicial (dd/mm/aaaa)", "data"),
    ("numero_folhas_autos_defesa_previa", "Nº de Folhas dos Autos (Vista Inicial)", "texto"),
    ("numero_folha_inicial_defesa_previa", "Nº da Folha Inicial (Vista Inicial)", "texto"),
    ("numero_folha_final_defesa_previa", "Nº da Folha Final (Vista Inicial)", "texto"),
    ("data_defesa_previa_apresentada", "Data de Apresentação da Defesa Prévia (dd/mm/aaaa)", "data"),
    # Vista Final (RED)
    ("data_vista_final", "Data de Abertura da Vista Final/RED (dd/mm/aaaa)", "data"),
    ("numero_folhas_autos_red", "Nº de Folhas dos Autos (RED)", "texto"),
    ("numero_folha_inicial_red", "Nº da Folha Inicial (RED)", "texto"),
    ("numero_folha_final_red", "Nº da Folha Final (RED)", "texto"),
    ("data_red_apresentada", "Data de Apresentação da RED (dd/mm/aaaa)", "data"),
    # Os modelos de Vista Inicial/Final pedem o inciso e o artigo do CEDM
    # separados (não uma citação única como no Despacho/Relatório)
    ("numero_inciso_cedm", "Nº do Inciso do CEDM (Vista Inicial/Final)", "texto"),
    ("numero_artigo_cedm", "Nº do Artigo do CEDM (Vista Inicial/Final)", "texto"),
    # Oitiva (notificação de testemunha e do sindicado/defensor)
    ("data_oitiva", "Data da Oitiva (dd/mm/aaaa)", "data"),
    ("hora_oitiva", "Hora da Oitiva", "texto"),
    ("endereco_sede", "Endereço da Sede da Unidade", "texto"),
    ("data_notificacao_testemunha", "Data de Notificação da Testemunha (dd/mm/aaaa)", "data"),
    ("data_notificacao_sindicado", "Data de Notificação do Sindicado p/ Audição (dd/mm/aaaa)", "data"),
    # Termo de Depoimento (qualificação completa da testemunha)
    ("numero_ordem_testemunha", "Número de Ordem da Testemunha", "texto"),
    ("nome_pai_testemunha", "Nome do Pai da Testemunha", "texto"),
    ("nome_mae_testemunha", "Nome da Mãe da Testemunha", "texto"),
    ("idade_testemunha", "Idade da Testemunha", "texto"),
    ("data_nascimento_testemunha", "Data de Nascimento da Testemunha (dd/mm/aaaa)", "data"),
    ("sexo_testemunha", "Sexo da Testemunha", "texto"),
    ("nacionalidade_testemunha", "Nacionalidade da Testemunha", "texto"),
    ("naturalidade_testemunha", "Naturalidade da Testemunha", "texto"),
    ("estado_civil_testemunha", "Estado Civil da Testemunha", "texto"),
    ("cpf_testemunha", "CPF da Testemunha", "texto"),
    ("identidade_testemunha", "Identidade da Testemunha", "texto"),
    ("local_trabalho_testemunha", "Local de Trabalho da Testemunha", "texto"),
    ("telefone_celular_testemunha", "Telefone Celular da Testemunha", "texto"),
    ("telefone_residencial_testemunha", "Telefone Residencial da Testemunha", "texto"),
    ("telefone_comercial_testemunha", "Telefone Comercial da Testemunha", "texto"),
    ("escolaridade_testemunha", "Escolaridade da Testemunha", "texto"),
    ("teor_depoimento", "Teor do Depoimento", "texto"),
    ("hora_inicio_depoimento", "Hora de Início do Depoimento", "texto"),
    ("hora_fim_depoimento", "Hora de Encerramento do Depoimento", "texto"),
    ("nome_defensor_sindicado", "Nome do Defensor do Sindicado", "texto"),
    ("data_proxima_oitiva", "Data da Próxima Oitiva (se houver)", "texto"),
    ("hora_proxima_oitiva", "Hora da Próxima Oitiva (se houver)", "texto"),
    ("local_proxima_oitiva", "Local da Próxima Oitiva (se houver)", "texto"),
    # Relatório do Encarregado
    ("data_hora_militar_fato", "Data/Hora do Fato (formato interno da unidade)", "texto"),
    ("em_servico_sindicado", "Sindicado Estava em Serviço? (S/N)", "sim_nao"),
    ("numero_folha_depoimento_testemunha", "Nº da Folha do Depoimento da Testemunha", "texto"),
    ("objetos_apreendidos", "Objetos Apreendidos/Arrecadados", "texto"),
    ("outras_provas", "Outras Provas", "texto"),
    ("analise_fatos_e_provas", "Análise dos Fatos e das Provas (seção 2)", "texto"),
    ("alegacoes_defesa_analise", "Análise das Alegações de Defesa (seção 3)", "texto"),
    ("incidentes_processuais", "Incidentes Processuais (seção 4)", "texto"),
    ("data_relatorio", "Data do Relatório (dd/mm/aaaa)", "data"),
    # Ofício de Remessa
    ("numero_oficio_remessa", "Nº do Ofício de Remessa", "texto"),
    ("data_oficio_remessa", "Data do Ofício de Remessa (dd/mm/aaaa)", "data"),
    ("numero_folhas_autos_final", "Nº Total de Folhas dos Autos (Remessa)", "texto"),
    # Ata de Reunião do CEDMU (só exigida se houve RED - art. 523 do MAPPA)
    ("referencia_procedimento", "Referência do Procedimento (Despacho/Portaria nº)", "texto"),
    ("data_reuniao", "Data da Reunião do CEDMU (dd/mm/aaaa)", "data"),
    ("cidade_reuniao", "Cidade da Reunião do CEDMU", "texto"),
    ("local_reuniao", "Local da Reunião do CEDMU", "texto"),
    ("numero_conselho", "Nº do Conselho (CEDMU)", "texto"),
    ("numero_bie_conselho", "Nº do BI de Designação do Conselho", "texto"),
    ("data_bie_conselho", "Data do BI de Designação do Conselho (dd/mm/aaaa)", "data"),
    ("nome_presidente", "Nome do Presidente do Conselho", "texto"),
    ("posto_presidente", "Posto/Graduação do Presidente do Conselho", "texto"),
    ("re_presidente", "Número de Matrícula do Presidente do Conselho", "texto"),
    ("nome_membro", "Nome do Membro do Conselho", "texto"),
    ("posto_membro", "Posto/Graduação do Membro do Conselho", "texto"),
    ("re_membro", "Número de Matrícula do Membro do Conselho", "texto"),
    ("nome_escrivao", "Nome do Membro/Escrivão do Conselho", "texto"),
    ("posto_escrivao", "Posto/Graduação do Membro/Escrivão do Conselho", "texto"),
    ("re_escrivao", "Número de Matrícula do Membro/Escrivão do Conselho", "texto"),
    ("acusado_compareceu", "Acusado Compareceu à Reunião? (S/N)", "sim_nao"),
    ("qualificacao_acusado", "Qualificação do Acusado (nº, posto, nome)", "texto"),
    ("finalidade_texto", "Finalidade da Análise (seção 2 da ata)", "texto"),
    ("verificacao_preliminar_texto", "Verificação Preliminar (seção 3 da ata)", "texto"),
    ("fundamentacao_fatica_texto", "Fundamentação Fática (seção 4 da ata)", "texto"),
    ("fundamentacao_legal_texto", "Fundamentação Legal (seção 5 da ata)", "texto"),
    ("analise_merito_texto", "Análise de Mérito (seção 6 da ata)", "texto"),
    ("parecer_texto", "Parecer do CEDMU (seção 7 da ata)", "texto"),
    ("hora_inicio_reuniao", "Hora de Início da Reunião do CEDMU", "texto"),
    ("hora_fim_reuniao", "Hora de Encerramento da Reunião do CEDMU", "texto"),
]

COLUNA_STATUS = "Status"
COLUNA_MENSAGEM = "Mensagem do Sistema"
COLUNA_STATUS_VISTA_INICIAL = "Status Vista Inicial"
COLUNA_MENSAGEM_VISTA_INICIAL = "Mensagem Vista Inicial"
COLUNA_STATUS_OITIVA = "Status Oitiva"
COLUNA_MENSAGEM_OITIVA = "Mensagem Oitiva"
COLUNA_STATUS_DEPOIMENTO = "Status Depoimento"
COLUNA_MENSAGEM_DEPOIMENTO = "Mensagem Depoimento"
COLUNA_STATUS_VISTA_FINAL = "Status Vista Final"
COLUNA_MENSAGEM_VISTA_FINAL = "Mensagem Vista Final"
COLUNA_STATUS_RELATORIO = "Status Relatório"
COLUNA_MENSAGEM_RELATORIO = "Mensagem Relatório"
COLUNA_STATUS_OFICIO = "Status Ofício"
COLUNA_MENSAGEM_OFICIO = "Mensagem Ofício"
COLUNA_STATUS_CEDMU = "Status CEDMU"
COLUNA_MENSAGEM_CEDMU = "Mensagem CEDMU"
CABECALHOS = [c[1] for c in DEFINICAO_COLUNAS] + [
    COLUNA_STATUS, COLUNA_MENSAGEM,
    COLUNA_STATUS_VISTA_INICIAL, COLUNA_MENSAGEM_VISTA_INICIAL,
    COLUNA_STATUS_OITIVA, COLUNA_MENSAGEM_OITIVA,
    COLUNA_STATUS_DEPOIMENTO, COLUNA_MENSAGEM_DEPOIMENTO,
    COLUNA_STATUS_VISTA_FINAL, COLUNA_MENSAGEM_VISTA_FINAL,
    COLUNA_STATUS_RELATORIO, COLUNA_MENSAGEM_RELATORIO,
    COLUNA_STATUS_OFICIO, COLUNA_MENSAGEM_OFICIO,
    COLUNA_STATUS_CEDMU, COLUNA_MENSAGEM_CEDMU,
]


def _parse_sim_nao(valor) -> bool | None:
    if valor is None or str(valor).strip() == "":
        return None
    v = str(valor).strip().upper()
    if v in ("S", "SIM", "TRUE", "1"):
        return True
    if v in ("N", "NAO", "NÃO", "FALSE", "0"):
        return False
    raise ValueError(f"Valor S/N inválido: {valor!r}")


def _parse_data(valor) -> date | None:
    if valor is None or str(valor).strip() == "":
        return None
    if isinstance(valor, datetime):
        return valor.date()
    if isinstance(valor, date):
        return valor
    return datetime.strptime(str(valor).strip(), "%d/%m/%Y").date()


_PARSERS = {"texto": lambda v: (str(v).strip() if v not in (None, "") else None),
            "data": _parse_data,
            "sim_nao": _parse_sim_nao}

# Lookup público (chave do schema -> (rótulo, tipo)) - reaproveitado pelo
# assistente interativo (`pcd_automation.interativo`) para não duplicar a
# definição de rótulo/tipo de cada campo.
CAMPOS_INFO: dict[str, tuple[str, str]] = {chave: (rotulo, tipo) for chave, rotulo, tipo in DEFINICAO_COLUNAS}


def converter_valor(tipo: str, valor_bruto):
    """Converte um valor bruto (string digitada, célula de planilha) para o
    tipo Python correspondente - mesma lógica usada ao ler a planilha."""
    return _PARSERS[tipo](valor_bruto)


@dataclass
class LinhaProcesso:
    numero_linha: int  # linha na planilha (1-indexed, para escrever de volta)
    dados: dict
    status_atual: str | None


def criar_planilha_modelo(caminho: Path | str) -> Path:
    caminho = Path(caminho)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Processos"
    ws.append(CABECALHOS)
    for col_idx in range(1, len(CABECALHOS) + 1):
        cel = ws.cell(row=1, column=col_idx)
        cel.font = Font(bold=True)
        ws.column_dimensions[get_column_letter(col_idx)].width = 28
    ws.freeze_panes = "A2"
    caminho.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(caminho))
    return caminho


def _mapear_cabecalhos(ws) -> dict:
    cabecalho_para_indice = {cel.value: idx for idx, cel in enumerate(ws[1], start=1)}
    faltando = [c[1] for c in DEFINICAO_COLUNAS if c[1] not in cabecalho_para_indice]
    if faltando:
        raise ValueError(f"Colunas ausentes na planilha: {faltando}")
    return cabecalho_para_indice


def _ler_dados_linha(ws, cabecalho_para_indice: dict, linha_num: int) -> dict:
    dados: dict = {}
    for chave, cabecalho, tipo in DEFINICAO_COLUNAS:
        col_idx = cabecalho_para_indice[cabecalho]
        valor_bruto = ws.cell(row=linha_num, column=col_idx).value
        dados[chave] = _PARSERS[tipo](valor_bruto)
    return dados


def _linha_vazia(ws, linha_num: int) -> bool:
    valores = [ws.cell(row=linha_num, column=col).value for col in range(1, ws.max_column + 1)]
    return all(v in (None, "") for v in valores)


def ler_processos(caminho: Path | str, apenas_pendentes: bool = True) -> list[LinhaProcesso]:
    """Lê a planilha e retorna uma linha por processo (etapa de instauração).

    Se `apenas_pendentes`, ignora linhas cujo Status já não esteja vazio
    (já processadas anteriormente), evitando reinstaurar por engano.
    """
    wb = openpyxl.load_workbook(str(caminho), data_only=True)
    ws = wb.active
    cabecalho_para_indice = _mapear_cabecalhos(ws)
    status_idx = cabecalho_para_indice.get(COLUNA_STATUS)

    linhas: list[LinhaProcesso] = []
    for linha_num in range(2, ws.max_row + 1):
        if _linha_vazia(ws, linha_num):
            continue

        status_atual = ws.cell(row=linha_num, column=status_idx).value if status_idx else None
        if apenas_pendentes and status_atual:
            continue

        dados = _ler_dados_linha(ws, cabecalho_para_indice, linha_num)
        linhas.append(LinhaProcesso(numero_linha=linha_num, dados=dados, status_atual=status_atual))

    return linhas


def _ler_pendentes_de_etapa(caminho: Path | str, coluna_status_etapa: str) -> list[LinhaProcesso]:
    """Lê linhas já instauradas (Status=INSTAURADO) cuja etapa indicada por
    `coluna_status_etapa` ainda não foi realizada (coluna vazia)."""
    wb = openpyxl.load_workbook(str(caminho), data_only=True)
    ws = wb.active
    cabecalho_para_indice = _mapear_cabecalhos(ws)
    status_idx = cabecalho_para_indice[COLUNA_STATUS]
    status_etapa_idx = cabecalho_para_indice[coluna_status_etapa]

    linhas: list[LinhaProcesso] = []
    for linha_num in range(2, ws.max_row + 1):
        if _linha_vazia(ws, linha_num):
            continue

        status_atual = ws.cell(row=linha_num, column=status_idx).value
        status_etapa_atual = ws.cell(row=linha_num, column=status_etapa_idx).value
        if status_atual != "INSTAURADO" or status_etapa_atual:
            continue

        dados = _ler_dados_linha(ws, cabecalho_para_indice, linha_num)
        linhas.append(LinhaProcesso(numero_linha=linha_num, dados=dados, status_atual=status_atual))

    return linhas


def ler_pendentes_vista_inicial(caminho: Path | str) -> list[LinhaProcesso]:
    return _ler_pendentes_de_etapa(caminho, COLUNA_STATUS_VISTA_INICIAL)


def ler_pendentes_oitiva(caminho: Path | str) -> list[LinhaProcesso]:
    return _ler_pendentes_de_etapa(caminho, COLUNA_STATUS_OITIVA)


def ler_pendentes_depoimento(caminho: Path | str) -> list[LinhaProcesso]:
    return _ler_pendentes_de_etapa(caminho, COLUNA_STATUS_DEPOIMENTO)


def ler_pendentes_vista_final(caminho: Path | str) -> list[LinhaProcesso]:
    return _ler_pendentes_de_etapa(caminho, COLUNA_STATUS_VISTA_FINAL)


def ler_pendentes_relatorio(caminho: Path | str) -> list[LinhaProcesso]:
    return _ler_pendentes_de_etapa(caminho, COLUNA_STATUS_RELATORIO)


def ler_pendentes_oficio(caminho: Path | str) -> list[LinhaProcesso]:
    return _ler_pendentes_de_etapa(caminho, COLUNA_STATUS_OFICIO)


def ler_pendentes_cedmu(caminho: Path | str) -> list[LinhaProcesso]:
    """Linhas instauradas, com RED apresentada e sem a Ata do CEDMU ainda.

    Diferente das demais etapas: CEDMU só é obrigatório quando há RED final
    apresentada (art. 523, §§1º-2º, do MAPPA) - linhas sem
    `data_red_apresentada` são deliberadamente excluídas, não é uma
    pendência a resolver.
    """
    return [
        linha for linha in _ler_pendentes_de_etapa(caminho, COLUNA_STATUS_CEDMU)
        if linha.dados.get("data_red_apresentada")
    ]


def _atualizar_coluna_status(
    caminho: Path | str, numero_linha: int, coluna_status: str, coluna_mensagem: str, status: str, mensagem: str
) -> None:
    wb = openpyxl.load_workbook(str(caminho))
    ws = wb.active
    cabecalho_para_indice = {cel.value: idx for idx, cel in enumerate(ws[1], start=1)}
    ws.cell(row=numero_linha, column=cabecalho_para_indice[coluna_status]).value = status
    ws.cell(row=numero_linha, column=cabecalho_para_indice[coluna_mensagem]).value = mensagem
    wb.save(str(caminho))


def atualizar_status(caminho: Path | str, numero_linha: int, status: str, mensagem: str) -> None:
    _atualizar_coluna_status(caminho, numero_linha, COLUNA_STATUS, COLUNA_MENSAGEM, status, mensagem)


def atualizar_status_vista_inicial(caminho: Path | str, numero_linha: int, status: str, mensagem: str) -> None:
    _atualizar_coluna_status(
        caminho, numero_linha, COLUNA_STATUS_VISTA_INICIAL, COLUNA_MENSAGEM_VISTA_INICIAL, status, mensagem
    )


def atualizar_status_oitiva(caminho: Path | str, numero_linha: int, status: str, mensagem: str) -> None:
    _atualizar_coluna_status(
        caminho, numero_linha, COLUNA_STATUS_OITIVA, COLUNA_MENSAGEM_OITIVA, status, mensagem
    )


def atualizar_status_depoimento(caminho: Path | str, numero_linha: int, status: str, mensagem: str) -> None:
    _atualizar_coluna_status(
        caminho, numero_linha, COLUNA_STATUS_DEPOIMENTO, COLUNA_MENSAGEM_DEPOIMENTO, status, mensagem
    )


def atualizar_status_vista_final(caminho: Path | str, numero_linha: int, status: str, mensagem: str) -> None:
    _atualizar_coluna_status(
        caminho, numero_linha, COLUNA_STATUS_VISTA_FINAL, COLUNA_MENSAGEM_VISTA_FINAL, status, mensagem
    )


def atualizar_status_relatorio(caminho: Path | str, numero_linha: int, status: str, mensagem: str) -> None:
    _atualizar_coluna_status(
        caminho, numero_linha, COLUNA_STATUS_RELATORIO, COLUNA_MENSAGEM_RELATORIO, status, mensagem
    )


def atualizar_status_oficio(caminho: Path | str, numero_linha: int, status: str, mensagem: str) -> None:
    _atualizar_coluna_status(
        caminho, numero_linha, COLUNA_STATUS_OFICIO, COLUNA_MENSAGEM_OFICIO, status, mensagem
    )


def atualizar_status_cedmu(caminho: Path | str, numero_linha: int, status: str, mensagem: str) -> None:
    _atualizar_coluna_status(
        caminho, numero_linha, COLUNA_STATUS_CEDMU, COLUNA_MENSAGEM_CEDMU, status, mensagem
    )
